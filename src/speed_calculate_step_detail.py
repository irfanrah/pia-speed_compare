"""Per-step latency breakdown for one PE production tick (B=42) and one
FT_PE T=3 production tick (B=54). Mirrors the row layout of the
``Stage detail`` tables in ``README.md`` — every row is a discrete step
inside ``_preprocess_stage`` / ``_inference_stage`` / ``_postprocess_stage``,
and the ``in_full / in_half / in_inference / in_input_gen / in_cos_sim``
columns mark which stage each step belongs to.

The five stage totals at the top of each sheet are **Excel SUMPRODUCT
formulas** over the inclusion columns times ``mean_ms`` — not values
computed in Python. Open the workbook and you can verify or re-aggregate
without rerunning.

Output: ``results/step_detail_<gpu>_<timestamp>.xlsx`` with one sheet per
model. CLI knobs: ``--pe-batch 42 --ftpe-batch 54 --frames 3``.
"""

from __future__ import annotations

import argparse
import os
import statistics
import subprocess
import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import numpy as np
import torch
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / "src" / "Product-AI-mono" / "packages"))

DEFAULT_PE_ENGINE = REPO_ROOT / "assets" / "model" / "PE-Core-L14-336.engine"
DEFAULT_FTPE_ENGINE = REPO_ROOT / "assets" / "model" / "FT_PE-Core-L14-336_260318_vision_no_mean_pooling.engine"
DEFAULT_PE_TXT = REPO_ROOT / "assets" / "model" / "text_features.json"
DEFAULT_FTPE_TXT = REPO_ROOT / "assets" / "model" / "FT_text_features.json"
DEFAULT_RESULTS_DIR = REPO_ROOT / "results"

# ---- Env-var injection BEFORE importing services ----------------------------
_pre = argparse.ArgumentParser(add_help=False)
_pre.add_argument("--pe-engine", type=Path, default=DEFAULT_PE_ENGINE)
_pre.add_argument("--pe-text-features", type=Path, default=DEFAULT_PE_TXT)
_pre.add_argument("--ftpe-engine", type=Path, default=DEFAULT_FTPE_ENGINE)
_pre.add_argument("--ftpe-text-features", type=Path, default=DEFAULT_FTPE_TXT)
_pre_args, _ = _pre.parse_known_args()
os.environ["MODEL_PERCEPTION_ENCODER_TRT_PATH"] = str(_pre_args.pe_engine)
os.environ["MODEL_PERCEPTION_ENCODER_TXT_FEATURE_PATH"] = str(_pre_args.pe_text_features)
os.environ["MODEL_PE_VIOLENCE_DETECTION_TRT_PATH"] = str(_pre_args.ftpe_engine)
os.environ["FT_PE_TEXT_FEATURES_JSON"] = str(_pre_args.ftpe_text_features)
os.environ.setdefault("FT_PE_MODE", "8fps")
# -----------------------------------------------------------------------------

from queue import Queue  # noqa: E402

from pia.ai.tasks.T2VRet.models.PE.utils.complexity_check import (  # noqa: E402
    cuda_sync,
)
from pia.vision.preprocessing import cv_bgr2rgb_batch  # noqa: E402
from pia_prod.AI.modules.perception_encoder.service import PEService  # noqa: E402
from pia_prod.AI.modules.perception_encoder.trt_utils import (  # noqa: E402
    preprocess_image as pe_preprocess_image,
)
from pia_prod.AI.modules.perception_encoder.config import (  # noqa: E402
    TEMPORAL_SIZE as PE_TEMPORAL_SIZE,
)
from pia_prod.AI.modules.ft_pe.service import FTPEService  # noqa: E402
from pia_prod.AI.modules.ft_pe.config import (  # noqa: E402
    ABNORMAL_CLASS_NAMES,
    TEMPORAL_SIZE as FTPE_TEMPORAL_SIZE,
)
from pia_prod.AI.global_config import USER_PARAM_KEY, RET_EVENT_KEY  # noqa: E402


_FRAME_RNG = np.random.default_rng(0)
_FRAME_HW = (1080, 1920, 3)


def gen_random_frame() -> np.ndarray:
    return _FRAME_RNG.integers(0, 256, size=_FRAME_HW, dtype=np.uint8)


def time_step(fn):
    """Sync GPU, time the call, sync again. Unified for CPU+GPU work."""
    cuda_sync()
    t0 = time.perf_counter()
    out = fn()
    cuda_sync()
    return out, (time.perf_counter() - t0) * 1000.0


def make_pe_user_params(count: int) -> list[dict]:
    return [
        {"user_param": {
            "retEvent": {
                "fire_ret":     {"roi": {"polygonCoordinates": []}},
                "falldown_ret": {"roi": {"polygonCoordinates": []}},
                "smoke_ret":    {"roi": {"polygonCoordinates": []}},
            },
            "cameraId": f"cam_{i}",
            "organization": "pia",
        }}
        for i in range(count)
    ]


def make_ftpe_user_params(count: int) -> list[dict]:
    return [
        {"user_param": {
            "retEvent": {
                "fire_ft_ret":     {"roi": {"polygonCoordinates": []}},
                "falldown_ft_ret": {"roi": {"polygonCoordinates": []}},
                "smoke_ft_ret":    {"roi": {"polygonCoordinates": []}},
            },
            "cameraId": f"cam_{i}",
            "organization": "pia",
        }}
        for i in range(count)
    ]


# ============================================================================
# PE per-step timer
# ============================================================================

# Step layout — order must match README "Stage detail" PE table.
# (label, in_full, in_half, in_inference, in_input_gen, in_cos_sim)
PE_STEP_DEFS: list[tuple[str, int, int, int, int, int]] = [
    ("gen_random_frame x B",                        0, 0, 0, 1, 0),
    ("fresh_batches() (copy from pool)",            1, 1, 0, 0, 0),
    ("x = preprocessed tensor reused (no work)",    0, 0, 1, 0, 0),
    ("_preprocess_stage (cv_bgr2rgb+ROI+resize)",   1, 1, 0, 0, 0),
    ("_inference_stage TRT (B,1024)",               1, 1, 1, 0, 0),
    ("deque append per stream",                     1, 0, 0, 0, 0),
    ("mean-pool + cos-sim (mean@gpu_vectors.T)",    1, 0, 0, 0, 1),
    ("_decide_top_category_opt (topK=13)",          1, 0, 0, 0, 0),
    ("process_category",                            1, 0, 0, 0, 0),
    ("check_alarm_duration",                        1, 0, 0, 0, 0),
]


def bench_pe(service: PEService, batch_size: int, warmup: int, iters: int) -> dict[str, list[float]]:
    # Pre-generate disjoint batches for the entire run (warmup + measure).
    pool = [
        [gen_random_frame() for _ in range(batch_size)]
        for _ in range(warmup + iters)
    ]
    stream_ids = [f"stream_{i}" for i in range(batch_size)]
    user_params = make_pe_user_params(batch_size)

    # ---- Warmup: drive full prod tick to warm the pipeline + kernels ----
    for j in range(warmup):
        batches = [arr.copy() for arr in pool[j]]
        service._detect(batches=batches, stream_ids=stream_ids, user_params=user_params)

    samples: dict[str, list[float]] = defaultdict(list)

    for j in range(iters):
        slot = warmup + j

        # 1. gen_random_frame x B
        _, ms = time_step(lambda: [gen_random_frame() for _ in range(batch_size)])
        samples[PE_STEP_DEFS[0][0]].append(ms)

        # 2. fresh_batches() — .copy() from pre-gen pool
        (batches,), ms = time_step(lambda: ([arr.copy() for arr in pool[slot]],))
        samples[PE_STEP_DEFS[1][0]].append(ms)

        # 3. x = preprocessed tensor reused — pure plumbing, zero-cost label
        samples[PE_STEP_DEFS[2][0]].append(0.0)

        # 4. _preprocess_stage  (BGR→RGB + ROI crop + resize + normalize)
        def step_preprocess():
            if not service.is_torch_batches(batches, speed_mode=True):
                cv_bgr2rgb_batch(batches)
            cropped = service.roi_manager.process_batches_with_roi(batches, user_params)
            return pe_preprocess_image(cropped)
        x, ms = time_step(step_preprocess)
        samples[PE_STEP_DEFS[3][0]].append(ms)

        # 5. _inference_stage  → (B, 1024)
        visual_vectors, ms = time_step(lambda: service.model(x))
        samples[PE_STEP_DEFS[4][0]].append(ms)

        # 6. deque append per stream  (stream_vector_queues maxlen=TEMPORAL_SIZE=1)
        def step_deque():
            for sid, vv in zip(stream_ids, visual_vectors):
                while service.stream_vector_queues[sid].__len__() < PE_TEMPORAL_SIZE:
                    service.stream_vector_queues[sid].append(service.zero_mask_vec)
                service.stream_vector_queues[sid].append(vv)
        _, ms = time_step(step_deque)
        samples[PE_STEP_DEFS[5][0]].append(ms)

        # 7. mean-pool + cos-sim   (per-stream sum/len @ gpu_vectors.T)
        evt = service.alarm_event_manager
        sims: list[torch.Tensor] = []
        def step_meancos():
            sims.clear()
            for sid in stream_ids:
                vq = service.stream_vector_queues[sid]
                m = (sum(vq) / len(vq))[None, :]
                sims.append((m @ evt.gpu_vectors.T).squeeze(dim=0))
        _, ms = time_step(step_meancos)
        samples[PE_STEP_DEFS[6][0]].append(ms)

        # 8. _decide_top_category_opt   (topK=13 ranking)
        predict_list: list = []
        def step_topk():
            predict_list.clear()
            for sim in sims:
                preds, _info = evt._decide_top_category_opt(sim)
                predict_list.append(preds)
        _, ms = time_step(step_topk)
        samples[PE_STEP_DEFS[7][0]].append(ms)

        # 9. process_category
        def step_process():
            for sid, user_param, preds in zip(stream_ids, user_params, predict_list):
                evt.process_category(user_param, preds, sid)
        _, ms = time_step(step_process)
        samples[PE_STEP_DEFS[8][0]].append(ms)

        # 10. check_alarm_duration
        _, ms = time_step(evt.check_alarm_duration)
        samples[PE_STEP_DEFS[9][0]].append(ms)

    return samples


# ============================================================================
# FT_PE per-step timer
# ============================================================================

FTPE_STEP_DEFS: list[tuple[str, int, int, int, int, int]] = [
    ("gen_random_frame x B",                                  0, 0, 0, 1, 0),
    ("fresh_batches() (copy from pool)",                      1, 1, 0, 0, 0),
    ("expand to (B,T,3,H,W) — reused (no work)",              0, 0, 1, 0, 0),
    ("_preprocess_stage (cv_bgr2rgb+ROI+resize)",             1, 1, 0, 0, 0),
    ("gather_frame_buffers append + _unconsumed++",           1, 1, 0, 0, 0),
    ("window check + torch.stack -> (B_enc,T,3,H,W)",         1, 1, 0, 0, 0),
    ("_inference_stage TRT + L2 norm (B_enc,T,1024)",         1, 1, 1, 0, 0),
    ("frame_buffers extend + rotate (del [1] if >TEMP)",      1, 1, 0, 0, 0),
    ("per-stream torch.stack(buf).mean(dim=0)",               1, 1, 0, 0, 0),
    ("torch.stack video_emb -> (B_ready,1024)",               1, 1, 0, 0, 0),
    ("L2 norm of video embeddings",                           1, 0, 0, 0, 0),
    ("per-class cos-sim (vis@cat_txt[c]).max(1)",             1, 0, 0, 0, 1),
    ("(sim_abn_max > sim_nrm_max).cpu().tolist()",            1, 0, 0, 0, 0),
    ("alarm_event_manager.update",                            1, 0, 0, 0, 0),
]


def bench_ftpe(service: FTPEService, batch_size: int, frames: int,
               warmup: int, iters: int) -> dict[str, list[float]]:
    # Stride=1 by overriding service window/sliding/prediction so every
    # measure tick triggers an encode + decision.
    service.window_size = frames
    service.sliding_window_size = max(0, frames - 1)
    service.prediction_size = 1
    service.stride = max(1, service.window_size - service.sliding_window_size)

    # Buffer warmup ticks must be enough to fill TEMPORAL_SIZE video embeddings
    # AND prime the rotating frame_buffer del[1] path.
    buffer_warmup = FTPE_TEMPORAL_SIZE + frames + 2
    total_pool = buffer_warmup + warmup + iters

    pool = [
        [gen_random_frame() for _ in range(batch_size)]
        for _ in range(total_pool)
    ]
    stream_ids = [f"stream_{i}" for i in range(batch_size)]
    user_params = make_ftpe_user_params(batch_size)

    # Drive `_detect` ticks during buffer + extra warmup so the postprocess
    # encode + decision path is hot before measure starts.
    for j in range(buffer_warmup + warmup):
        batches = [arr.copy() for arr in pool[j]]
        service._detect(batches=batches, stream_ids=stream_ids, user_params=user_params)

    samples: dict[str, list[float]] = defaultdict(list)

    for j in range(iters):
        slot = buffer_warmup + warmup + j

        # 1. gen_random_frame x B
        _, ms = time_step(lambda: [gen_random_frame() for _ in range(batch_size)])
        samples[FTPE_STEP_DEFS[0][0]].append(ms)

        # 2. fresh_batches()
        (batches,), ms = time_step(lambda: ([arr.copy() for arr in pool[slot]],))
        samples[FTPE_STEP_DEFS[1][0]].append(ms)

        # 3. expand to 5D — plumbing label
        samples[FTPE_STEP_DEFS[2][0]].append(0.0)

        # 4. _preprocess_stage
        processed_batches, ms = time_step(lambda: service._preprocess_stage(batches, user_params))
        samples[FTPE_STEP_DEFS[3][0]].append(ms)

        # 5. gather_frame_buffers append + _unconsumed_frames++
        def step_gather_append():
            for sid, b in zip(stream_ids, processed_batches):
                service.gather_frame_buffers[sid].append(b)
                service._unconsumed_frames[sid] += 1
        _, ms = time_step(step_gather_append)
        samples[FTPE_STEP_DEFS[4][0]].append(ms)

        # 6. window check + torch.stack(gather) → (B_enc, T, 3, H, W)
        encode_stream_ids: list[str] = []
        encode_tensors: list[torch.Tensor] = []
        def step_window():
            encode_stream_ids.clear()
            encode_tensors.clear()
            for sid in stream_ids:
                gbuf = service.gather_frame_buffers[sid]
                if (len(gbuf) == service.window_size
                        and service._unconsumed_frames[sid] >= service.stride):
                    encode_tensors.append(torch.stack(list(gbuf)))
                    service._unconsumed_frames[sid] -= service.stride
                    encode_stream_ids.append(sid)
            if encode_stream_ids:
                # build the encode-batch tensor too — same shape the engine sees
                stacked = torch.stack(encode_tensors)
                encode_tensors.clear()
                encode_tensors.append(stacked)
        _, ms = time_step(step_window)
        samples[FTPE_STEP_DEFS[5][0]].append(ms)

        # 7. _inference_stage (model fwd) + L2 norm
        embeddings: list = []  # mutable slot
        def step_infer():
            embeddings.clear()
            if encode_stream_ids:
                embeddings.append(service._inference_stage(encode_tensors[0]))
        _, ms = time_step(step_infer)
        samples[FTPE_STEP_DEFS[6][0]].append(ms)

        # 8. frame_buffers extend + rotate
        def step_fb_rotate():
            if not encode_stream_ids:
                return
            for sid, emb in zip(encode_stream_ids, embeddings[0]):
                new_embs = emb[-service.prediction_size:]
                service.frame_buffers[sid].extend(new_embs.unbind(0))
                if len(service.frame_buffers[sid]) > FTPE_TEMPORAL_SIZE:
                    del service.frame_buffers[sid][1]
        _, ms = time_step(step_fb_rotate)
        samples[FTPE_STEP_DEFS[7][0]].append(ms)

        # 9. per-stream torch.stack(buf).mean(dim=0)
        ready_stream_ids: list[str] = []
        video_embeddings: list[torch.Tensor] = []
        def step_meanpool():
            ready_stream_ids.clear()
            video_embeddings.clear()
            for sid in stream_ids:
                buf = service.frame_buffers[sid]
                if len(buf) >= FTPE_TEMPORAL_SIZE:
                    stacked = torch.stack(list(buf))
                    video_embeddings.append(stacked.mean(dim=0))
                    ready_stream_ids.append(sid)
        _, ms = time_step(step_meanpool)
        samples[FTPE_STEP_DEFS[8][0]].append(ms)

        # 10. torch.stack video_emb → (B_ready, 1024)
        vis_box: list[torch.Tensor] = []
        def step_stack_vid():
            vis_box.clear()
            if video_embeddings:
                vis_box.append(torch.stack(video_embeddings))
        _, ms = time_step(step_stack_vid)
        samples[FTPE_STEP_DEFS[9][0]].append(ms)

        # 11. L2 norm of (B, 1024) video embeddings
        def step_l2():
            if vis_box:
                v = vis_box[0]
                vis_box[0] = v / v.norm(dim=-1, keepdim=True)
        _, ms = time_step(step_l2)
        samples[FTPE_STEP_DEFS[10][0]].append(ms)

        # 12. per-class cos-sim   (vis @ cat_txt[c]).max(1) for each class
        cossim_box: dict[str, tuple[torch.Tensor, torch.Tensor]] = {}
        def step_cossim():
            cossim_box.clear()
            if not vis_box:
                return
            vis = vis_box[0]
            for class_name in ABNORMAL_CLASS_NAMES:
                txt_vec = service.category_txt_vectors.get(class_name)
                normal_vec = service.category_normal_vectors.get(class_name)
                if txt_vec is None or normal_vec is None:
                    continue
                sim_abn_max = (vis @ txt_vec).max(dim=1).values
                sim_nrm_max = (vis @ normal_vec).max(dim=1).values
                cossim_box[class_name] = (sim_abn_max, sim_nrm_max)
        _, ms = time_step(step_cossim)
        samples[FTPE_STEP_DEFS[11][0]].append(ms)

        # 13. (sim_abn_max > sim_nrm_max).cpu().tolist()
        category_preds: dict[str, list[bool]] = {}
        def step_decision():
            category_preds.clear()
            for class_name, (abn, nrm) in cossim_box.items():
                category_preds[class_name] = (abn > nrm).detach().cpu().tolist()
        _, ms = time_step(step_decision)
        samples[FTPE_STEP_DEFS[12][0]].append(ms)

        # 14. alarm_event_manager.update
        def step_alarm():
            if not ready_stream_ids:
                return
            preds_per_stream = [
                {cls: category_preds[cls][i] for cls in category_preds}
                for i in range(len(ready_stream_ids))
            ]
            user_param_list = [user_params[stream_ids.index(sid)] for sid in ready_stream_ids]
            service.alarm_event_manager.update(preds_per_stream, ready_stream_ids, user_param_list)
        _, ms = time_step(step_alarm)
        samples[FTPE_STEP_DEFS[13][0]].append(ms)

    return samples


# ============================================================================
# Excel writer
# ============================================================================

def write_sheet(wb: Workbook, sheet_name: str, header_label: str,
                step_defs: list[tuple[str, int, int, int, int, int]],
                samples: dict[str, list[float]]) -> None:
    """Layout per sheet:

        Row 1  : title
        Row 3-7: stage totals (SUMPRODUCT formulas over step rows)
        Row 9  : table header
        Row 10+: one row per step
    """
    ws = wb.create_sheet(sheet_name)
    bold = Font(bold=True)
    fill_total = PatternFill("solid", fgColor="DCE6F1")
    fill_hdr   = PatternFill("solid", fgColor="305496")
    fill_hdr_font = Font(bold=True, color="FFFFFF")

    ws["A1"] = header_label
    ws["A1"].font = Font(bold=True, size=12)

    # Reserve rows 3..7 for cycle totals (formulas filled after we know row range)
    totals = [
        ("full_cycle (ms)",      "in_full"),
        ("half_cycle (ms)",      "in_half"),
        ("inference (ms)",       "in_inference"),
        ("input_gen (ms)",       "in_input_gen"),
        ("cos_sim (ms)",         "in_cos_sim"),
    ]
    for i, (label, _flag) in enumerate(totals):
        r = 3 + i
        ws.cell(row=r, column=1, value=label).font = bold
        ws.cell(row=r, column=1).fill = fill_total
        # value cell filled below once table range is known
        ws.cell(row=r, column=2).fill = fill_total

    header_row = 9
    columns = [
        "step", "in_full", "in_half", "in_inference", "in_input_gen", "in_cos_sim",
        "mean_ms", "std_ms", "min_ms", "max_ms", "iters",
    ]
    for ci, name in enumerate(columns, start=1):
        c = ws.cell(row=header_row, column=ci, value=name)
        c.font = fill_hdr_font
        c.fill = fill_hdr
        c.alignment = Alignment(horizontal="center")

    # Step rows
    for ri, (label, in_full, in_half, in_inf, in_inp, in_cos) in enumerate(step_defs):
        r = header_row + 1 + ri
        ms_list = samples.get(label, [])
        if ms_list:
            mean = round(statistics.fmean(ms_list), 4)
            std = round(statistics.pstdev(ms_list), 4) if len(ms_list) > 1 else 0.0
            mn = round(min(ms_list), 4)
            mx = round(max(ms_list), 4)
            n = len(ms_list)
        else:
            mean = std = mn = mx = 0.0
            n = 0
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=in_full)
        ws.cell(row=r, column=3, value=in_half)
        ws.cell(row=r, column=4, value=in_inf)
        ws.cell(row=r, column=5, value=in_inp)
        ws.cell(row=r, column=6, value=in_cos)
        ws.cell(row=r, column=7, value=mean)
        ws.cell(row=r, column=8, value=std)
        ws.cell(row=r, column=9, value=mn)
        ws.cell(row=r, column=10, value=mx)
        ws.cell(row=r, column=11, value=n)

    first_row = header_row + 1
    last_row = header_row + len(step_defs)
    flag_cols = {"in_full": "B", "in_half": "C", "in_inference": "D",
                 "in_input_gen": "E", "in_cos_sim": "F"}
    mean_range = f"G{first_row}:G{last_row}"
    for i, (label, flag_name) in enumerate(totals):
        r = 3 + i
        col = flag_cols[flag_name]
        flag_range = f"{col}{first_row}:{col}{last_row}"
        formula = f"=SUMPRODUCT({flag_range},{mean_range})"
        ws.cell(row=r, column=2, value=formula).font = bold

    # Column widths
    widths = [54, 8, 8, 13, 13, 11, 11, 11, 11, 11, 7]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def gpu_info_str(device_index: int = 0) -> str:
    try:
        n = torch.cuda.get_device_name(device_index)
    except Exception:
        n = "unknown"
    return n


# ============================================================================
# Main
# ============================================================================

def main() -> int:
    p = argparse.ArgumentParser(parents=[_pre])
    p.add_argument("--pe-batch", type=int, default=42)
    p.add_argument("--ftpe-batch", type=int, default=54)
    p.add_argument("--frames", type=int, default=3, help="T for FT_PE")
    p.add_argument("--warmup", type=int, default=5)
    p.add_argument("--iters", type=int, default=25)
    p.add_argument("--device-index", type=int, default=0)
    p.add_argument("--out-dir", type=Path, default=DEFAULT_RESULTS_DIR)
    args = p.parse_args()

    if not torch.cuda.is_available():
        print("CUDA not available", file=sys.stderr)
        return 2

    args.out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    gpu_tag = gpu_info_str(args.device_index).replace(" ", "_")
    out_xlsx = args.out_dir / f"step_detail_{gpu_tag}_{ts}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    # ---- PE ----
    print(f"\n=== PE  B={args.pe_batch}  warmup={args.warmup}  iters={args.iters} ===")
    pe_service = PEService(Queue())
    pe_samples = bench_pe(pe_service, args.pe_batch, args.warmup, args.iters)
    for label, ms_list in pe_samples.items():
        if ms_list:
            print(f"  {label:50s}  mean={statistics.fmean(ms_list):8.3f} ms")
    write_sheet(wb, "PE_B42",
                header_label=(f"PE  |  B={args.pe_batch}  T=1  "
                              f"warmup={args.warmup}  iters={args.iters}  "
                              f"|  GPU: {gpu_info_str(args.device_index)}"),
                step_defs=PE_STEP_DEFS, samples=pe_samples)
    del pe_service
    torch.cuda.empty_cache()

    # ---- FT_PE ----
    print(f"\n=== FT_PE B={args.ftpe_batch} T={args.frames} "
          f"warmup={args.warmup} iters={args.iters} ===")
    ftpe_service = FTPEService(Queue())
    ftpe_samples = bench_ftpe(ftpe_service, args.ftpe_batch, args.frames,
                              args.warmup, args.iters)
    for label, ms_list in ftpe_samples.items():
        if ms_list:
            print(f"  {label:50s}  mean={statistics.fmean(ms_list):8.3f} ms")
    write_sheet(wb, "FTPE_B54_T3",
                header_label=(f"FT_PE  |  B={args.ftpe_batch}  T={args.frames}  "
                              f"warmup={args.warmup}  iters={args.iters}  "
                              f"|  GPU: {gpu_info_str(args.device_index)}"),
                step_defs=FTPE_STEP_DEFS, samples=ftpe_samples)

    wb.save(out_xlsx)
    print(f"\nwrote: {out_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
